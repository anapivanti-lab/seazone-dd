"""Controle de Due Diligences INTEGRADO ao sistema: a cada parecer gerado, o
sistema registra/atualiza a DD num cadastro (JSON persistente na pasta das
Franquias). A página /controle mostra todas as DDs numa tabela. Há também
exportação para Excel sob demanda (botão), mas o controle vive no próprio sistema.
"""
from __future__ import annotations

import html
import json
import re
import uuid
from datetime import datetime

from .config import base_saida

ARQ_REG = "_controle_dds.json"
# Campos editáveis de cada DD (na ordem da tabela)
CAMPOS = ["id", "franquia", "cnpj", "representante", "cpf", "cidade_uf", "risco", "link", "obs", "data"]
_COR = {"ALTO": ("#c0392b", "#f8d7da"), "MÉDIO": ("#b8860b", "#fff3cd"), "BAIXO": ("#1a7d3c", "#d4edda")}


def _fmt(doc: str, pj: bool) -> str:
    d = re.sub(r"\D", "", doc or "")
    if pj and len(d) == 14:
        return f"{d[:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:]}"
    if not pj and len(d) == 11:
        return f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}"
    return doc or ""


def _caminho():
    return base_saida() / ARQ_REG


def _salvar(regs: list) -> bool:
    try:
        _caminho().write_text(json.dumps(regs, ensure_ascii=False, indent=2), encoding="utf-8")
        return True
    except Exception:
        return False


def _novo_rid() -> str:
    return uuid.uuid4().hex[:12]


def listar() -> list:
    """Lê o cadastro. Garante que todo registro tenha um 'rid' (id estável usado para
    editar/excluir mesmo depois de reordenar a lista) — se faltar, cria e regrava."""
    try:
        regs = json.loads(_caminho().read_text(encoding="utf-8"))
    except Exception:
        return []
    mudou = False
    for r in regs:
        if not r.get("rid"):
            r["rid"] = _novo_rid()
            mudou = True
    if mudou:
        _salvar(regs)
    return regs


def salvar_registro(dados: dict) -> dict:
    """Cria (sem rid) ou edita (com rid) uma DD a partir dos campos da tela."""
    regs = listar()
    rid = str(dados.get("rid", "") or "").strip()
    rec = {k: str(dados.get(k, "") or "").strip() for k in CAMPOS}
    if rid:
        for r in regs:
            if r.get("rid") == rid:
                r.update(rec)
                _salvar(regs)
                return r
    # novo registro
    rec["rid"] = _novo_rid()
    if not rec["data"]:
        rec["data"] = datetime.now().strftime("%d/%m/%Y")
    regs.append(rec)
    _salvar(regs)
    return rec


def excluir_registro(rid: str) -> bool:
    """Exclui a DD pelo rid. Devolve True se removeu algo."""
    regs = listar()
    novo = [r for r in regs if r.get("rid") != str(rid)]
    if len(novo) == len(regs):
        return False
    return _salvar(novo)


def _record(d: dict, url_pasta: str) -> dict:
    ents = d.get("entidades", [])
    pj = next((e for e in ents if e.get("tipo") == "PJ"), None)
    pf = (next((e for e in ents if "operador" in e.get("papel", "").lower()), None)
          or next((e for e in ents if e.get("tipo") == "PF"), None))
    cidade = next((f"{e.get('municipio', '')}/{e.get('uf', '')}".strip("/")
                   for e in ents if e.get("municipio") or e.get("uf")), "")
    obs = d.get("concl_texto", "")
    if d.get("recomendacoes"):
        obs += " Recomendações: " + "; ".join(d["recomendacoes"])
    return {
        "id": str(d.get("id_suporte", "") or ""),
        "franquia": (pj.get("nome") if pj else (pf.get("nome") if pf else d.get("titulo"))) or "—",
        "cnpj": _fmt(pj.get("documento", ""), True) if pj else "",
        "representante": pf.get("nome", "") if pf else "",
        "cpf": _fmt(pf.get("documento", ""), False) if pf else "",
        "cidade_uf": cidade, "risco": d.get("risco", ""), "link": url_pasta or "",
        "obs": obs, "data": datetime.now().strftime("%d/%m/%Y"),
    }


def registrar(d: dict, url_pasta: str = "") -> dict:
    """Cria/atualiza a DD no cadastro (chaveia pelo ID Suporte; senão CNPJ+CPF)."""
    rec = _record(d, url_pasta)
    regs = listar()
    idx = None
    for i, r in enumerate(regs):
        if rec["id"] and r.get("id") == rec["id"]:
            idx = i
            break
        if not rec["id"] and (rec["cnpj"] or rec["cpf"]) and (r.get("cnpj"), r.get("cpf")) == (rec["cnpj"], rec["cpf"]):
            idx = i
            break
    if idx is None:
        rec["rid"] = _novo_rid()
        regs.append(rec)
    else:
        rec["rid"] = regs[idx].get("rid") or _novo_rid()  # mantém o id estável ao atualizar
        regs[idx] = rec
    _salvar(regs)
    return rec


def importar_planilha(caminho: str) -> int:
    """Importa as DDs já feitas de uma planilha (xlsx) para o cadastro do sistema.
    Colunas: Franquia | CNPJ | Representante | CPF | Cidade/UF | Risco | Link | Obs.
    Não duplica (chaveia por CNPJ+CPF). Devolve quantas foram adicionadas."""
    from openpyxl import load_workbook
    wb = load_workbook(caminho)
    ws = wb.active
    regs = listar()
    chaves = {(r.get("cnpj"), r.get("cpf")) for r in regs if (r.get("cnpj") or r.get("cpf"))}
    add = 0
    for row in ws.iter_rows(min_row=2):
        v = [(c.value if c.value is not None else "") for c in row[:8]]
        while len(v) < 8:
            v.append("")
        franquia, cnpj, rep, cpf, cidade, risco, _link, obs = [str(x).strip() for x in v]
        link = (row[6].hyperlink.target if (len(row) > 6 and row[6].hyperlink) else "") or ""
        if not any([franquia, cnpj, rep, cpf, obs]):
            continue
        chave = (cnpj or None, cpf or None)
        if any(chave) and chave in chaves:
            continue
        regs.append({"id": "", "franquia": franquia, "cnpj": cnpj, "representante": rep, "cpf": cpf,
                     "cidade_uf": cidade, "risco": risco, "link": link, "obs": obs, "data": ""})
        if any(chave):
            chaves.add(chave)
        add += 1
    _salvar(regs)
    return add


# ---------------------------------------------------------------- página /controle
_ESTILO = """
*{box-sizing:border-box}
body{font-family:'Segoe UI',Arial,sans-serif;color:#1a2332;margin:0;background:#eef1f4}
header{background:#1a2b3c;color:#fff;padding:1rem 2rem;display:flex;justify-content:space-between;align-items:center;gap:1rem;flex-wrap:wrap}
header h1{margin:0;font-size:1.2rem}
header a{color:#9fd0e8;text-decoration:none;font-size:.9rem}
.btn{background:#1a7d3c;color:#fff;padding:.5rem 1rem;border-radius:8px;text-decoration:none;font-weight:600;font-size:.85rem}
main{max-width:1320px;margin:1.3rem auto;padding:0 1rem}
.resumo{display:flex;gap:.7rem;margin-bottom:1rem;flex-wrap:wrap}
.pill{background:#fff;border-radius:10px;padding:.5rem 1rem;box-shadow:0 1px 4px rgba(0,0,0,.07);font-size:.8rem;color:#5a6b7a;text-align:center}
.pill b{font-size:1.25rem;display:block;color:#1a2332}
.busca{padding:.6rem .9rem;border:1px solid #ccd5dd;border-radius:9px;font-size:.95rem;width:340px;max-width:100%;margin-bottom:.9rem}
.tabela{background:#fff;border-radius:12px;box-shadow:0 1px 6px rgba(0,0,0,.08);overflow:hidden}
table{border-collapse:collapse;width:100%;table-layout:fixed}
th,td{padding:.55rem .7rem;text-align:left;font-size:.82rem;vertical-align:top;border-bottom:1px solid #eef1f4;overflow:hidden}
th{background:#0b4f6c;color:#fff;font-weight:600}
tr:hover td{background:#f6f9fb}
.fr{font-weight:600}
.muted{color:#8a97a3;font-size:.78rem}
.doc{font-variant-numeric:tabular-nums;white-space:nowrap}
.risco{display:inline-block;font-weight:700;text-align:center;border-radius:999px;padding:.12rem .6rem;font-size:.74rem}
td a{color:#0b6;font-weight:600;text-decoration:none;white-space:nowrap}
.clamp2{display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical;overflow:hidden}
.obs .txt{display:-webkit-box;-webkit-line-clamp:3;-webkit-box-orient:vertical;overflow:hidden;cursor:pointer}
.obs .txt.aberto{-webkit-line-clamp:unset}
.vazio{background:#fff;padding:2rem;text-align:center;border-radius:12px;color:#777}
.barra{display:flex;justify-content:space-between;align-items:center;gap:1rem;flex-wrap:wrap}
.btn-add{background:#1a7d3c;color:#fff;border:none;padding:.55rem 1rem;border-radius:8px;font-weight:700;font-size:.85rem;cursor:pointer}
.acao{cursor:pointer;border:none;background:none;font-size:1rem;padding:.05rem .2rem;line-height:1}
.acao:hover{transform:scale(1.15)}
/* modal de edição/criação */
.modal{position:fixed;inset:0;background:rgba(0,0,0,.45);display:none;align-items:flex-start;justify-content:center;padding:2rem 1rem;overflow:auto;z-index:50}
.modal.aberto{display:flex}
.card{background:#fff;border-radius:14px;padding:1.4rem 1.5rem;max-width:600px;width:100%;box-shadow:0 10px 40px rgba(0,0,0,.3)}
.card h2{margin:.1rem 0 1rem;font-size:1.15rem;color:#1a2b3c}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:.7rem}
.campo{margin-bottom:.6rem}
.campo.full{grid-column:1/-1}
.campo label{display:block;font-size:.76rem;color:#5a6b7a;margin-bottom:.2rem;font-weight:600}
.campo input,.campo select,.campo textarea{width:100%;padding:.5rem .6rem;border:1px solid #ccd5dd;border-radius:8px;font-size:.9rem;font-family:inherit}
.campo textarea{min-height:72px;resize:vertical}
.acoes-modal{display:flex;justify-content:flex-end;gap:.6rem;margin-top:1.1rem}
.btn-cancel{background:#eef1f4;color:#33414f;border:none;padding:.55rem 1.1rem;border-radius:8px;font-weight:600;cursor:pointer}
.btn-salvar{background:#1a7d3c;color:#fff;border:none;padding:.55rem 1.3rem;border-radius:8px;font-weight:700;cursor:pointer}
.btn-salvar:disabled{opacity:.6;cursor:default}
"""

_COLGROUP = ("<colgroup>"
             '<col style="width:50px"><col style="width:15%"><col style="width:128px"><col style="width:13%">'
             '<col style="width:108px"><col style="width:96px"><col style="width:80px"><col style="width:58px">'
             '<col><col style="width:74px"><col style="width:70px"></colgroup>')


def pagina_html() -> str:
    regs = listar()
    n = len(regs)
    por_risco = {}
    for r in regs:
        rr = (r.get("risco") or "").upper()
        if rr in ("ALTO", "MÉDIO", "BAIXO"):
            por_risco[rr] = por_risco.get(rr, 0) + 1
    pills = f'<div class="pill">Total<b>{n}</b></div>'
    for risco in ("ALTO", "MÉDIO", "BAIXO"):
        if por_risco.get(risco):
            cf, cb = _COR[risco]
            pills += f'<div class="pill" style="background:{cb};color:{cf}">{risco.title()}<b>{por_risco[risco]}</b></div>'

    e = html.escape
    # dados crus (não formatados pra exibição) usados pelo formulário de edição
    dados_js = json.dumps({r["rid"]: r for r in regs}, ensure_ascii=False).replace("</", "<\\/")

    barra = ('<div class="barra">'
             '<input class="busca" id="q" placeholder="🔎 Filtrar por nome, CNPJ, cidade, risco…" onkeyup="filtra()">'
             '<button class="btn-add" onclick="nova()">➕ Adicionar DD</button></div>')

    if not regs:
        corpo = (barra + '<div class="vazio">Nenhuma DD registrada ainda. '
                 'Clique em “➕ Adicionar DD” ou gere um parecer que ela aparece aqui.</div>')
    else:
        linhas = ""
        for r in reversed(regs):  # mais recentes primeiro
            rid = e(r.get("rid", ""))
            risco = (r.get("risco") or "").strip()
            cf, cb = _COR.get(risco.upper(), ("#5a6b7a", "#eef1f4"))
            badge = (f'<span class="risco" style="color:{cf};background:{cb}">{e(risco)}</span>'
                     if risco else '<span class="muted">—</span>')
            link = f'<a href="{e(r["link"])}" target="_blank">abrir ↗</a>' if r.get("link") else "—"
            obs = e(r.get("obs", ""))
            linhas += (
                f'<tr><td class="muted">{e(r.get("id", ""))}</td>'
                f'<td><div class="fr clamp2" title="{e(r.get("franquia", ""))}">{e(r.get("franquia", ""))}</div></td>'
                f'<td class="doc">{e(r.get("cnpj", ""))}</td>'
                f'<td><div class="clamp2">{e(r.get("representante", ""))}</div></td>'
                f'<td class="doc">{e(r.get("cpf", ""))}</td>'
                f'<td>{e(r.get("cidade_uf", ""))}</td>'
                f'<td>{badge}</td><td>{link}</td>'
                f'<td class="obs"><div class="txt" title="{obs}" onclick="this.classList.toggle(\'aberto\')">{obs}</div></td>'
                f'<td class="muted">{e(r.get("data", ""))}</td>'
                f'<td style="white-space:nowrap">'
                f'<button class="acao" title="Editar" onclick="editar(\'{rid}\')">✏️</button>'
                f'<button class="acao" title="Excluir" onclick="excluir(\'{rid}\')">🗑️</button></td></tr>')
        corpo = (
            barra +
            f'<div class="tabela"><table id="tab">{_COLGROUP}'
            '<thead><tr><th>ID</th><th>Franquia</th><th>CNPJ</th><th>Representante / Operador</th>'
            '<th>CPF</th><th>Cidade/UF</th><th>Risco</th><th>Pasta</th><th>Observações</th><th>Data</th>'
            '<th>Ações</th></tr></thead>'
            f'<tbody>{linhas}</tbody></table></div>')

    modal = (
        '<div class="modal" id="modal"><div class="card">'
        '<h2 id="modal_titulo">Editar DD</h2><input type="hidden" id="f_rid">'
        '<div class="grid2">'
        '<div class="campo"><label>ID Suporte (Pipefy)</label><input id="f_id"></div>'
        '<div class="campo"><label>Data</label><input id="f_data" placeholder="dd/mm/aaaa"></div>'
        '<div class="campo full"><label>Franquia (razão social)</label><input id="f_franquia"></div>'
        '<div class="campo"><label>CNPJ</label><input id="f_cnpj"></div>'
        '<div class="campo"><label>Cidade/UF</label><input id="f_cidade_uf"></div>'
        '<div class="campo full"><label>Representante / Operador</label><input id="f_representante"></div>'
        '<div class="campo"><label>CPF</label><input id="f_cpf"></div>'
        '<div class="campo"><label>Risco</label><select id="f_risco">'
        '<option value="">—</option><option value="BAIXO">BAIXO</option>'
        '<option value="MÉDIO">MÉDIO</option><option value="ALTO">ALTO</option></select></div>'
        '<div class="campo full"><label>Link da pasta (Drive)</label><input id="f_link" placeholder="https://..."></div>'
        '<div class="campo full"><label>Observações</label><textarea id="f_obs"></textarea></div>'
        '</div>'
        '<div class="acoes-modal"><button class="btn-cancel" onclick="fecharModal()">Cancelar</button>'
        '<button class="btn-salvar" id="btnSalvar" onclick="salvar()">Salvar</button></div>'
        '</div></div>')

    script = (
        '<script>'
        f'const REGS={dados_js};'
        'const CAMPOS=["rid","id","franquia","cnpj","representante","cpf","cidade_uf","risco","link","obs","data"];'
        'function filtra(){var q=document.getElementById("q").value.toLowerCase();'
        'document.querySelectorAll("#tab tbody tr").forEach(function(tr){'
        'tr.style.display=tr.innerText.toLowerCase().includes(q)?"":"none";});}'
        'function gv(id){return document.getElementById(id);}'
        'function sv(id,v){gv(id).value=(v==null?"":v);}'
        'function abrir(){gv("modal").classList.add("aberto");}'
        'function fecharModal(){gv("modal").classList.remove("aberto");}'
        'function nova(){CAMPOS.forEach(function(c){sv("f_"+c,"");});'
        'gv("modal_titulo").textContent="Adicionar DD";abrir();}'
        'function editar(rid){var r=REGS[rid]||{};'
        'sv("f_rid",rid);sv("f_id",r.id);sv("f_franquia",r.franquia);sv("f_cnpj",r.cnpj);'
        'sv("f_representante",r.representante);sv("f_cpf",r.cpf);sv("f_cidade_uf",r.cidade_uf);'
        'sv("f_risco",(r.risco||"").toUpperCase());sv("f_link",r.link);sv("f_obs",r.obs);sv("f_data",r.data);'
        'gv("modal_titulo").textContent="Editar DD";abrir();}'
        'async function salvar(){var body={};CAMPOS.forEach(function(c){body[c]=gv("f_"+c).value;});'
        'var b=gv("btnSalvar");b.disabled=true;b.textContent="Salvando…";'
        'try{var r=await fetch("/controle/salvar",{method:"POST",'
        'headers:{"Content-Type":"application/json"},body:JSON.stringify(body)});'
        'if(!r.ok)throw 0;location.reload();}'
        'catch(e){b.disabled=false;b.textContent="Salvar";alert("Não consegui salvar. Tente de novo.");}}'
        'async function excluir(rid){var r=REGS[rid]||{};'
        'if(!confirm("Excluir a DD \\""+(r.franquia||"")+"\\"? Isso não pode ser desfeito."))return;'
        'try{var resp=await fetch("/controle/excluir",{method:"POST",'
        'headers:{"Content-Type":"application/json"},body:JSON.stringify({rid:rid})});'
        'if(!resp.ok)throw 0;location.reload();}'
        'catch(e){alert("Não consegui excluir. Tente de novo.");}}'
        'document.addEventListener("keydown",function(ev){if(ev.key==="Escape")fecharModal();});'
        '</script>')

    return (f'<!doctype html><html lang="pt-br"><head><meta charset="utf-8">'
            f'<meta name="viewport" content="width=device-width, initial-scale=1">'
            f'<title>Controle de Due Diligences</title><style>{_ESTILO}</style></head><body>'
            f'<header><h1>📊 Controle de Due Diligences</h1>'
            f'<span><a href="/">← Voltar ao sistema</a> &nbsp;&nbsp; '
            f'<a href="/controle.xlsx" class="btn">⬇️ Exportar Excel</a></span></header>'
            f'<main><div class="resumo">{pills}</div>{corpo}</main>{modal}{script}</body></html>')


def exportar_xlsx() -> str:
    """Gera o Excel sob demanda (a partir do cadastro) e devolve o caminho."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    cols = ["ID Suporte", "Franquia", "CNPJ", "Representante/Operador", "CPF", "Cidade/UF",
            "Risco", "Link da DD", "Observações", "Data"]
    larg = [11, 34, 20, 30, 16, 16, 9, 38, 70, 12]
    wb = Workbook()
    ws = wb.active
    ws.title = "Due Diligences"
    for c, t in enumerate(cols, 1):
        cel = ws.cell(1, c, t)
        cel.font = Font(bold=True, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor="0B4F6C")
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = larg[c - 1]
    ws.freeze_panes = "A2"
    for r in listar():
        ws.append([r.get("id", ""), r.get("franquia", ""), r.get("cnpj", ""), r.get("representante", ""),
                   r.get("cpf", ""), r.get("cidade_uf", ""), r.get("risco", ""), r.get("link", ""),
                   r.get("obs", ""), r.get("data", "")])
        if r.get("link"):
            cel = ws.cell(ws.max_row, 8)
            cel.value = "abrir pasta"
            cel.hyperlink = r["link"]
            cel.font = Font(color="0563C1", underline="single")
        ws.cell(ws.max_row, 9).alignment = Alignment(wrap_text=True, vertical="top")
    destino = base_saida() / "Controle de Due Diligence.xlsx"
    wb.save(str(destino))
    return str(destino)
